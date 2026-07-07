import csv
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _assert_row_limit(count: int, max_rows: int) -> None:
    if count > max_rows:
        raise HTTPException(status_code=400, detail=f"导入数据不能超过 {max_rows} 行")


def read_csv_bytes(content: bytes, *, max_rows: int) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV 文件必须使用 UTF-8 编码") from exc
    reader = csv.DictReader(StringIO(text))
    result: list[dict[str, str]] = []
    for index, row in enumerate(reader, start=1):
        _assert_row_limit(index, max_rows)
        result.append(dict(row))
    return result


def read_xlsx_bytes(content: bytes, *, max_rows: int) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Excel 文件无法读取或格式不正确") from exc
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        headers_row = next(row_iter)
    except StopIteration:
        workbook.close()
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in headers_row]
    result: list[dict[str, Any]] = []
    for index, row in enumerate(row_iter, start=1):
        _assert_row_limit(index, max_rows)
        result.append({headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)})
    workbook.close()
    return result


def workbook_to_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
